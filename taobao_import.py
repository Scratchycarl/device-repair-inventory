"""Parse Taobao order exports and match rows to pending order-part jobs."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from repair_parts import normalize_model, screen_part_for_model

# (keywords, canonical part name). None = infer screen part from device model.
PART_KEYWORD_GROUPS: list[tuple[list[str], str | None]] = [
    (["激光雷达", "雷达", "lidar", "测距仪", "测距", "写码雷达"], "LiDAR Module"),
    (["尾插", "充电口", "充电接口", "数据线接口"], "Charging Port"),
    (["后盖", "后玻璃", "背玻璃", "背板", "后壳"], None),  # resolved per device
    (["电池", "电芯"], "Replacement Battery"),
    (["摄像头", "相机", "后置摄像", "前置摄像"], "Camera Module"),
    (["听筒"], "Earpiece Speaker"),
    (["扬声器", "喇叭"], "Loudspeaker"),
    (["wifi天线", "wifi 天线", "wifi"], "WiFi Antenna"),
    (["face id", "面容", "truedepth"], "Face ID / TrueDepth"),
    (["home键", "home button", "指纹键"], "Home Button"),
    (["oled", "屏幕总成", "显示屏", "屏幕", "液晶"], None),
    (["lcd"], "LCD Assembly"),
    (["digitizer", "触摸", "外屏"], "Digitizer"),
]


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "value"):
        value = value.value
        if value is None:
            return ""
    return str(value).strip()


def _resolve_col(col_map: dict[str, int], names: tuple[str, ...]) -> int | None:
    """Exact header first, then substring (Taobao adds parenthetical notes)."""
    for name in names:
        if name in col_map:
            return col_map[name]
    for header, idx in col_map.items():
        for name in names:
            if name and name in header:
                return idx
    return None


def _pick_column(cells: list, col_map: dict[str, int], names: tuple[str, ...]) -> str:
    idx = _resolve_col(col_map, names)
    if idx is None or idx >= len(cells):
        return ""
    return _cell_str(cells[idx])


def parse_taobao_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb.active
    # Materialize once — read-only sheets cannot be iterated twice.
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    wb.close()

    header_info = None
    for row_idx, row in enumerate(rows[:15], start=1):
        headers = [_cell_str(c) for c in row]
        if "订单号" in headers:
            col_map = {name: idx for idx, name in enumerate(headers) if name}
            header_info = (row_idx, col_map)
            break
    if not header_info:
        raise ValueError("Could not find Taobao header row (订单号).")

    header_row, col_map = header_info
    required = ["订单号", "商品名称", "型号款式"]
    for key in required:
        if key not in col_map:
            raise ValueError(f"Missing column: {key}")

    qty_col = _resolve_col(col_map, ("商品数量",))

    orders: list[dict[str, Any]] = []
    for row in rows[header_row:]:
        cells = list(row)
        order_id = _cell_str(cells[col_map["订单号"]]) if col_map["订单号"] < len(cells) else ""
        if not order_id:
            continue

        product_name = _cell_str(cells[col_map["商品名称"]]) if col_map["商品名称"] < len(cells) else ""
        variant = _cell_str(cells[col_map["型号款式"]]) if col_map["型号款式"] < len(cells) else ""
        qty = 1
        if qty_col is not None and qty_col < len(cells):
            raw_qty = cells[qty_col]
            try:
                qty = max(1, int(float(raw_qty)))
            except (TypeError, ValueError):
                qty = 1

        orders.append(
            {
                "order_id": order_id,
                "order_date": _cell_str(cells[col_map.get("订单提交时间", -1)])
                if col_map.get("订单提交时间") is not None
                and col_map["订单提交时间"] < len(cells)
                else "",
                "order_status": _cell_str(cells[col_map.get("订单状态", -1)])
                if col_map.get("订单状态") is not None
                and col_map["订单状态"] < len(cells)
                else "",
                "shop_name": _cell_str(cells[col_map.get("店铺名称", -1)])
                if col_map.get("店铺名称") is not None
                and col_map["店铺名称"] < len(cells)
                else "",
                "product_name": product_name,
                "variant": variant,
                "product_link": _cell_str(cells[col_map.get("商品链接", -1)])
                if col_map.get("商品链接") is not None
                and col_map["商品链接"] < len(cells)
                else "",
                "qty": qty,
                "domestic_carrier": _pick_column(cells, col_map, (
                    "物流公司", "快递公司", "承运公司",
                )),
                "domestic_tracking_number": _pick_column(cells, col_map, (
                    "物流单号", "运单号", "快递单号", "物流编号",
                )),
                "logistics_status": _pick_column(cells, col_map, (
                    "物流状态", "快递状态",
                )),
                "search_text": f"{product_name} {variant}".lower(),
            }
        )
    return orders


def _model_family(text: str) -> str | None:
    if "ipad" in text:
        return "ipad"
    if "iphone" in text:
        return "iphone"
    return None


def _model_flags(text: str) -> tuple[bool, bool, bool, bool]:
    return (
        "pro" in text,
        "max" in text or "plus" in text,
        "mini" in text,
        "air" in text,
    )


def extract_model_tokens(text: str) -> set[str]:
    """Pull an Apple generation from 型号款式 (e.g. '14', '15 pro max', '精诚13…')."""
    lowered = text.lower()
    tokens: set[str] = set()

    patterns = [
        r"iphone\s*\d{1,2}\s*pro\s*max",
        r"iphone\s*\d{1,2}\s*pro",
        r"iphone\s*\d{1,2}\s*plus",
        r"iphone\s*\d{1,2}\s*mini",
        r"iphone\s*\d{1,2}",
        r"ipad\s*pro",
        r"ipad\s*air",
        r"ipad\s*mini",
        r"ipad\s*\d+",
        r"(?<!\d)(1[0-7]|[6-9])(?!\d)\s*pro\s*max",
        r"(?<!\d)(1[0-7]|[6-9])(?!\d)\s*promax",
        r"(?<!\d)(1[0-7]|[6-9])(?!\d)\s*pro",
        r"(?<!\d)(1[0-7]|[6-9])(?!\d)\s*plus",
        r"(?<!\d)(1[0-7]|[6-9])(?!\d)\s*mini",
        r"(?<!\d)(1[0-7]|[6-9])(?!\d)",
        r"(?<![a-z0-9])se(?:\s*[123])?(?![a-z0-9])",
    ]
    for pattern in patterns:
        m = re.search(pattern, lowered)
        if m:
            tokens.add(normalize_model(m.group(0)))
            break
    return tokens


def order_model_tokens(order: dict[str, Any]) -> set[str]:
    """Model comes from 型号款式; fall back to 商品名称 if style has none."""
    tokens = extract_model_tokens(order.get("variant") or "")
    if tokens:
        return tokens
    return extract_model_tokens(order.get("product_name") or "")


def order_part_name(order: dict[str, Any], device_model: str | None = None) -> str | None:
    """Part type comes from 商品名称 keywords (电池, 屏幕, …)."""
    return infer_part_from_text(order.get("product_name") or "", device_model=device_model)


def models_compatible(device_model: str, order_tokens: set[str]) -> bool:
    device_norm = normalize_model(device_model)
    if not device_norm or not order_tokens:
        return False
    device_flags = _model_flags(device_norm)
    device_family = _model_family(device_norm)
    device_nums = re.findall(r"\d+", device_norm)
    for token in order_tokens:
        if not token:
            continue
        token_family = _model_family(token)
        if device_family and token_family and device_family != token_family:
            continue
        if _model_flags(token) != device_flags:
            continue
        token_nums = re.findall(r"\d+", token)
        if token == device_norm:
            return True
        if device_nums and token_nums and device_nums[0] == token_nums[0]:
            return True
        if not device_nums and not token_nums and device_family and device_family == token_family:
            return True
    return False


def infer_part_from_text(text: str, device_model: str | None = None) -> str | None:
    lowered = text.lower()
    for keywords, part_name in PART_KEYWORD_GROUPS:
        if any(kw.lower() in lowered for kw in keywords):
            if part_name is None:
                if any(k in lowered for k in ("后盖", "后玻璃", "背玻璃", "背板", "后壳")):
                    from repair_parts import back_part_for_model

                    return back_part_for_model(device_model)
                return screen_part_for_model(device_model)
            return part_name
    return None


def _part_names_match(job_part: str, inferred: str) -> bool:
    if not job_part or not inferred:
        return False
    if job_part.lower() == inferred.lower():
        return True
    if job_part.lower() in inferred.lower() or inferred.lower() in job_part.lower():
        return True
    return False
