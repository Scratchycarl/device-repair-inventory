"""Parse Taobao order-export xlsx files and upsert them into the purchases tables.

The exports come in three variants (paid / shipped / received) that share the
same base columns; the shipped variant adds logistics columns. Order-level
cells (order no, shop, totals, logistics) are only filled on the first row of
each order (merged-cell style), so they must be forward-filled.
"""

from __future__ import annotations

import re
from datetime import datetime, timezone

from openpyxl import load_workbook

# Header name -> internal field
HEADER_MAP = {
    "订单号": "order_no",
    "订单提交时间": "submit_time",
    "订单状态": "status_text",
    "店铺名称": "shop_name",
    "商品名称": "item_title",
    "商品链接": "item_link",
    "型号款式": "sku_text",
    "商品数量": "quantity",
    "商品金额": "unit_price",
    "实付金额": "order_total",
    "运费": "shipping_fee",
}
# Logistics headers include a parenthetical note; match by prefix.
HEADER_PREFIX_MAP = {
    "物流公司": "logistics_company",
    "物流单号": "tracking_no",
}

STATUS_MAP = {
    "买家已付款": "paid",
    "卖家已发货": "shipped",
    "交易成功": "received",
}
# Rank used so re-imports only ever advance an order's status.
STATUS_RANK = {"paid": 1, "shipped": 2, "received": 3}

# Order-level fields forward-filled from the first row of each order.
ORDER_FIELDS = (
    "order_no",
    "submit_time",
    "status_text",
    "shop_name",
    "order_total",
    "shipping_fee",
    "logistics_company",
    "tracking_no",
)


def _parse_price(value):
    if value is None:
        return None
    text = str(value).strip().replace("￥", "").replace("¥", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_qty(value):
    if value is None:
        return 1
    try:
        return max(1, int(float(str(value).strip())))
    except (ValueError, TypeError):
        return 1


def parse_taobao_xlsx(file_obj):
    """Parse an xlsx file object into a list of row dicts (one per item line).

    Raises ValueError if the sheet doesn't look like a Taobao order export.
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)

        header_row = next(rows_iter, None)
        if not header_row:
            raise ValueError("Empty sheet")

        col_fields = {}
        for idx, cell in enumerate(header_row):
            name = str(cell or "").strip()
            if not name:
                continue
            if name in HEADER_MAP:
                col_fields[idx] = HEADER_MAP[name]
            else:
                for prefix, field in HEADER_PREFIX_MAP.items():
                    if name.startswith(prefix):
                        col_fields[idx] = field
                        break

        required = {"order_no", "item_title", "sku_text"}
        found = set(col_fields.values())
        if not required.issubset(found):
            missing = required - found
            raise ValueError(
                f"Not a Taobao order export: missing columns {sorted(missing)}"
            )

        rows = []
        carry = {f: None for f in ORDER_FIELDS}
        for raw in rows_iter:
            record = {}
            for idx, field in col_fields.items():
                value = raw[idx] if idx < len(raw) else None
                if value is not None and str(value).strip() == "":
                    value = None
                record[field] = value

            # Forward-fill order-level fields from the order's first row.
            if record.get("order_no"):
                for f in ORDER_FIELDS:
                    carry[f] = record.get(f)
            else:
                for f in ORDER_FIELDS:
                    if record.get(f) is None:
                        record[f] = carry[f]

            if not record.get("order_no") or not record.get("item_title"):
                continue

            status_text = str(record.get("status_text") or "").strip()
            rows.append({
                "order_no": str(record["order_no"]).strip(),
                "submit_time": str(record.get("submit_time") or "").strip(),
                "status": STATUS_MAP.get(status_text, "paid"),
                "shop_name": str(record.get("shop_name") or "").strip(),
                "order_total": _parse_price(record.get("order_total")),
                "shipping_fee": _parse_price(record.get("shipping_fee")),
                "logistics_company": str(record.get("logistics_company") or "").strip() or None,
                "tracking_no": str(record.get("tracking_no") or "").strip() or None,
                "item_title": str(record["item_title"]).strip(),
                "item_link": str(record.get("item_link") or "").strip(),
                "sku_text": str(record.get("sku_text") or "").strip(),
                "quantity": _parse_qty(record.get("quantity")),
                "unit_price": _parse_price(record.get("unit_price")),
            })
        return rows
    finally:
        wb.close()


def upsert_purchases(conn, rows):
    """Idempotently upsert parsed rows. Returns counts dict.

    - New orders/items are inserted.
    - Existing orders only advance status (paid -> shipped -> received) and
      gain logistics info; they never regress.
    - Existing items are never duplicated (unique per order+title+sku).
    """
    counts = {"orders_new": 0, "orders_updated": 0, "items_new": 0, "items_skipped": 0}
    now = datetime.now(timezone.utc).isoformat()
    cur = conn.cursor()

    for row in rows:
        existing = cur.execute(
            "SELECT status, logistics_company, tracking_no FROM purchase_orders WHERE order_no = ?",
            (row["order_no"],),
        ).fetchone()

        if existing is None:
            cur.execute(
                """INSERT INTO purchase_orders
                   (order_no, submit_time, status, shop_name, order_total,
                    shipping_fee, logistics_company, tracking_no, imported_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    row["order_no"], row["submit_time"], row["status"],
                    row["shop_name"], row["order_total"], row["shipping_fee"],
                    row["logistics_company"], row["tracking_no"], now,
                ),
            )
            counts["orders_new"] += 1
        else:
            old_status = existing[0] or "paid"
            new_status = row["status"]
            updates = {}
            if STATUS_RANK.get(new_status, 0) > STATUS_RANK.get(old_status, 0):
                updates["status"] = new_status
            if row["logistics_company"] and not existing[1]:
                updates["logistics_company"] = row["logistics_company"]
            if row["tracking_no"] and not existing[2]:
                updates["tracking_no"] = row["tracking_no"]
            if updates:
                sets = ", ".join(f"{k} = ?" for k in updates)
                cur.execute(
                    f"UPDATE purchase_orders SET {sets} WHERE order_no = ?",
                    (*updates.values(), row["order_no"]),
                )
                counts["orders_updated"] += 1

        item_exists = cur.execute(
            "SELECT id FROM purchase_items WHERE order_no = ? AND item_title = ? AND sku_text = ?",
            (row["order_no"], row["item_title"], row["sku_text"]),
        ).fetchone()
        if item_exists:
            counts["items_skipped"] += 1
        else:
            cur.execute(
                """INSERT INTO purchase_items
                   (order_no, item_title, item_link, sku_text, quantity, unit_price)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (
                    row["order_no"], row["item_title"], row["item_link"],
                    row["sku_text"], row["quantity"], row["unit_price"],
                ),
            )
            counts["items_new"] += 1

    conn.commit()
    return counts
